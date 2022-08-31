import openpyxl
import os
import psycopg2
import configparser
import time
import re


def cod():
    global codec
    codec = ''
    codecs = ['cp1251', 'cp1252', 'utf-8']
    fig = configparser.ConfigParser()
    for j in codecs:
        try:
            fig.read('path.ini', encoding=j)
            directory = fig.get('con', 'directory') + '\\'  # директория где файлы лежат
            if 'Р' in directory or '°' in directory or '±' in directory or '†' in directory:
                continue
            else:
                codec = j
                break
        except UnicodeDecodeError:
            continue
    if codec == '':
        print('Не нашлось изестных кодировок. Принтскрин ошибки пришлите на электроннyю почту: ProhorenkoSV@nesk.ru')


cod()


def inidir():
    fig = configparser.ConfigParser()
    fig.read('path.ini', encoding=codec)
    directory = fig.get('con', 'directory') + '\\'  # директория где файлы лежат
    return directory  # результат считывания данных из ини-файла для последующего использования


def ini():
    fig = configparser.ConfigParser()
    fig.read('path.ini', encoding=codec)
    server = fig.get('base', 'server')  # наименование сервера
    port = fig.get('base', 'port')  # наименование базы
    base = fig.get('base', 'base')  # наименование базы
    return server, port, base  # результат считывания данных из ини-файла для последующего использования


def inicolumn():
    fig = configparser.ConfigParser()
    fig.read('path.ini', encoding=codec)
    datac = fig.get('column', 'data')  # № колонки с датой
    summac = fig.get('column', 'summa')  # № колонки с суммой
    commentc = fig.get('column', 'comment')  # № колонки с комментарием
    fil = fig.get('column', 'fil')  #
    return datac, summac, commentc, fil  # результат считывания данных из ини-файла для последующего использования


def connnect():
    server, port, base = ini()
    try:
        con = psycopg2.connect(dbname=base, user='SA', password='SA', host=server, port =port)
        return con
    except pyodbc.Error as err:
        print('Принтскрин ошибки пришлите на электроннyю почту: ProhorenkoSV@nesk.ru'+'\n' + err.args[1][50:])
        input('Нажмите Enter для выхода\n')


def pathf():
    directory = inidir()
    listpath = []
    count = 0  # кол-во файлов
    for dirpath, dirnames, filenames in os.walk(directory):
        for filename in filenames:
            if dirpath == directory and filename.startswith('Сбер') and filename.endswith('.xlsx'):
                g = str(directory + filename)
                print(dirpath + filename)
                listpath.append(g)
                count += 1
    dir2 = directory + 'out\\'
    if os.path.exists(dir2):
        print("Путь для переноса отработанных реестров существует!")
    else:
        os.mkdir(dir2)
        print("Создали папку для переноса отработанных реестров.")
    print("Кол-во файлов для обработки: " + str(count))
    print()
    return count, listpath, dir2


def obr():  # основная функция обработки файла
    tic = time.perf_counter()
    count, listpath, dir2 = pathf()
    datac, summac, commentc, fil = inicolumn()
    for k in range(0, count):
        con = connnect()
        list = []
        file = listpath[k].split('\\')[-1].strip()  # отсоединяю имя файла для последующего сохранения нового файла
        print(f'Читаю файл № {k+1}: ' + str(file))
        wookbook = openpyxl.load_workbook(listpath[k])  # открываю исходный эксель файл
        worksheet = wookbook.active
        a = worksheet.max_row
        # b = worksheet.max_column
        print(f'Колличество строк для обработки в файле № {k+1}: ' + str(a))
        summaob = 0  # объявляю переменную для общей суммы
        countob = 0  # счетчик кол-ва лицевых
        for row in range(1, a):
            summa = worksheet.cell(row, int(summac)).value  # сумма
            if summa:
                ost = re.findall('(?<![0-9])[0-9]{7,13}(?![0-9])', str((worksheet.cell(row, int(commentc))).value))  # комментарий
                data = (str((worksheet.cell(row, int(datac))).value)[:-9]).replace('.', '-')  # дата
                data = data.split('-')
                data.reverse()
                data = '-'.join(data)
                cursor = con.cursor()  # открываю курсор к базе
                for nls in ost:
                    sel = (
                        'select ls.row_id, ls."Номер", (\n'
                        '    case when at.Участок=\'Филиал АО «НЭСК» «Новороссийскэнергосбыт»\' then 958620\n'
                        '	when at.Участок=\'Филиал АО «НЭСК» «Туапсеэнергосбыт»\' then 644771\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Тимашевскэнергосбыт»\' then 690211\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Кропоткинэнергосбыт»\' then 776717\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Армавирэнергосбыт»\' then 516912\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Геленджикэнергосбыт»\' then 750668\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Горячеключэнергосбыт»\' then 668605\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Анапаэнергосбыт»\' then 983596\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Ейскэнергосбыт»\' then 577444\n'
                        '	 when at.Участок=\'Филиал АО «НЭСК» «Славянскэнергосбыт»\' then 859965\n'
                        '	 when at.Участок=\'Новокубанский участок\' then 476798\n'
                        '	 when at.Участок=\'Гулькевичский участок\' then 902305\n'
                        '	 when at.Участок=\'Приморско-Ахтарский участок\' then 838125\n'
                        '	 when at.Участок=\'Темрюкский участок\' then 877157\n'
                        '	 when at.Участок=\'Усть-Лабинский участок\' then 395398\n'
                        '	 when at.Участок=\'Западное отделение\' then 215788\n'
                        '	 when at.Участок=\'Центральное отделение\' then 291605\n'
                        '	 when at.Участок=\'Карасунское отделение\' then 372397\n'
                        '	 when at.Участок=\'Прикубанское отделение\' then 126156\n'
                        '	 when at.Участок=\'Апшеронский участок\' then 446018\n'
                        '	 when at.Участок=\'Лабинский участок\' then 416071\n'
                        '	 when at.Участок=\'Белореченский участок\' then 540209\n'
                        '	 when at.Участок=\'Тихорецкий участок\' then 820427\n'
                        '	 when at.Участок=\'Кореновский участок\' then 618515\n'
                        '	 when at.Участок=\'Крымский участок\' then 797437\n'
                        '	 when at.Участок=\'Абинский участок\' then 712138\n'
                        '	 when at.Участок=\'Курганинский участок\' then 598633\n'
                        '	 when at.Участок=\'Центр обслуживания потребителей Лабинского участка\' then 458916\n'
                        '	 end) as fil\n'
                        'from stack."Лицевые счета" ls\n'
                        'inner join lateral stack."AddrLs_Table"(ls.row_id, 1) at on true\n'
                        'where "Номер" = %s or "СтороннийНомер" = %s')
                    cursor.execute(sel, (nls, nls))  # выполняем запрос с параметрами
                    res = cursor.fetchone()  # полученный результат проверки лицевого
                    if res:
                        filial = res[2]
                        if int(fil) == int(filial):
                            row = res[0]  # преобразовываю в нужный формат
                            nls = res[1]
                            summa = round(summa * 100)
                            print('Строка номер:' + str(row) + '. Лицевой: ' + str(nls) + '. Сумма платежа: ' + str(summa/100))
                            summaob += summa
                            list.append([data, row, summa])
                            countob += 1
                            break
                cursor.close()  # закрываю курсор к базе
        count -= 1
        con.commit()  # закрываем подключение
        summaob = round(summaob/100, 2)  # округляю до двух знаков
        list.append(summaob)  # добавляем общую сумму платежей
        list.append(countob)  # добавляем кол-во платежей
        list.append(filial)  # добавляем филиал
        if len(list) < 4:
            print('Нечего обрабатывать :(')
        else:
            oblist, obrees, obdata, per = lis(list)  # список платежей для заголовка реестра
            telo(oblist, obrees, obdata, per)  # список данных для тела реестра
        print(f'Всего обработано лицевых: ' + str(countob) + ' на сумму ' + str(summaob))
        print(f'Файл № {k + 1} прочитан: ' + str(file))
        #os.replace(listpath[k], dir2+file)  # перенос файла
        print(f'Файл № {k + 1} перенесен в {dir2}.')
        print()
    print("Осталось считать строк: " + str(count))
    toc = time.perf_counter()
    print(f"Вычисление заняло {toc - tic:0.4f} секунд")
    input('Нажмите Enter для выхода\n')


def reestr():  # функция для получения id последнего реестра +1
    con = connnect()
    cur = con.cursor()
    sel = """select max("Номер")+1 from stack.Документ"""
    cur.execute(sel)
    rees = cur.fetchone()[0]
    cur.close()
    con.commit()
    return rees


def lis(list):  # функция создания заголовка реестра
    rees = reestr()
    dataob = list[0][0]  # дата заголовка и 3 строки преобразования
    dataob = dataob.split('-')
    dataob.reverse()
    per = dataob
    datazag = '-'.join(dataob) + ' 00:00:00'  # готовая дата заголовка
    per[2] = '01'
    per = '-'.join(per) + ' 00:00:00'  # рабочий период
    obsum = str(list[-3])
    kolvo = str(list[-2])  # кол-во платежей
    papka = str(list[-1])  # папка филиала
    con = connnect()
    curnom = con.cursor()
    sel = """select row_id from stack."Документ" where "Папки"=%s and "РасчМесяц" = %s and "Дата" = %s"""
    curnom.execute(sel, (papka, per, per))
    papres = curnom.fetchone()[0]
    curnom.close()
    cur = con.cursor()
    sel = (
        """insert into stack."Документ" (Дата,"Источник-Платежи","Кол_во","Номер","Папки", "Папки_add", "Примечание","РасчМесяц","Сумма","Тип документа")\n"""
        "	select %s Дата, -- дата реестра\n"
        "		475 , -- источник платежа в данном случае банки-прочие\n"
        "		%s, -- кол-во платежей\n"
        "		%s, -- номер реестра\n"
        "		%s, -- Папки - место где создаем заголовок реестра\n"
        "		1, -- Папки_add\n"
        "		1, -- примечание\n"
        "		%s, -- расчетный период\n"
        "		%s, -- сумма платежей\n"
        "		67 ")
    cur.execute(sel, (datazag, kolvo, rees, papres, per, obsum))
    cur.close()
    currees = con.cursor()
    sel = """select max(row_id) from stack."Документ" where "Папки" = %s"""
    currees.execute(sel, (papres,))
    paprees = currees.fetchone()[0]
    currees.close()
    con.commit()
    oblist = list
    obrees = paprees
    obdata = datazag
    return oblist, obrees, obdata, per


def telo(oblist, obrees, obdata, per):  # функция внутренности реестра
    con = connnect()
    nlscur = con.cursor()
    cur = con.cursor()
    for i in range(len(oblist)-3):
        nlsid = oblist[i][1]
        sumvrem = oblist[i][2]
        sel = (
            """insert into stack."Список оплаты" (Дата,"Пени","Платеж-Список","РасчМесяц","Сумма","Счет-Оплата")\n"""
            "	select %s Дата, -- дата реестра obdata\n"
            "		0, -- Пени \n"
            "		%s , -- Платеж-Список obrees\n"
            "		%s, -- РасчМесяц per\n"
            "		%s, -- Сумма платежа\n"
            "		%s -- Счет-Оплата лицевой")

        cur.execute(sel, (obdata, obrees, per, sumvrem, nlsid))
    nlscur.close()
    cur.close()
    con.commit()


obr()

