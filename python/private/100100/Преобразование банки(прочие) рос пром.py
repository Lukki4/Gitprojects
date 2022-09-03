import openpyxl
import os
import pyodbc
import configparser
import time
import re
#проверка создания ветки
#проверка после удаления ответвления

def cod():
    global codec
    codec = ''
    codecs = ['cp1251', 'cp1252', 'utf-8']
    fig = configparser.ConfigParser()
    for j in codecs:
        try:
            fig.read('path.ini', encoding=j)
            directory = fig.get('con', 'directory') + '\\'  # директория где файлы лежат
            if 'Р' in directory or '°' in directory or '±' in directory or '†' in directory:
                continue
            else:
                codec = j
                break
        except UnicodeDecodeError:
            continue
    if codec == '':
        print('Не нашлось изестных кодировок. Принтскрин ошибки пришлите на электроннyю почту: ProhorenkoSV@nesk.ru')


cod()


def inidir():
    fig = configparser.ConfigParser()
    fig.read('path.ini', encoding=codec)
    directory = fig.get('con', 'directory') + '\\'  # директория где файлы лежат
    return directory  # результат считывания данных из ини-файла для последующего использования


def ini():
    fig = configparser.ConfigParser()
    fig.read('path.ini', encoding=codec)
    driver = [x for x in pyodbc.drivers() if 'SQL Server' and '11' in x]
    if driver == []:
        driver = ['SQL Server']
    driver = driver[0]
    server1 = fig.get('base', 'server1')  # наименование сервера 1
    server2 = fig.get('base', 'server2')  # наименование сервера 2
    base1 = fig.get('base', 'base1')  # наименование базы 1
    base2 = fig.get('base', 'base2')  # наименование базы 2
    return driver, server1, server2, base1, base2  # результат считывания данных из ини-файла для последующего использования


def connnect():
    driver, server1, server2, base1, base2 = ini()
    try:
        conch = pyodbc.connect(f"""Driver={driver};
    Server={server1};Database={base1};UID=smisb_user;PWD=AaFDA2E2C77C""")
        conm = pyodbc.connect(f"""Driver={driver};
    Server={server2};Database={base2};UID=smisb_user;PWD=AaFDA2E2C77C""")
        return conch, conm
    except pyodbc.Error as err:
        print('Принтскрин ошибки пришлите на электроннyю почту: '+'\n' + err.args[1][50:])
        input('Нажмите Enter для выхода\n')


def pathf():
    directory = inidir()  # вызов результата ини-файла
    listpath = []
    count = 0  # кол-во файлов
    for dirpath, dirnames, filenames in os.walk(directory):
        for filename in filenames:
            if dirpath == directory and filename.startswith(('40702810203220000375', 'Statement')) and filename.endswith('.xlsx'):  # поиск только в текущей папке!
                g = str(directory + filename)
                print(dirpath + filename)
                listpath.append(g)
                count += 1
    dir2 = directory + 'out\\'
    if os.path.exists(dir2):
        print("Путь для переноса отработанных реестров существует!")
    else:
        os.mkdir(dir2)
        print("Создали папку для переноса отработанных реестров.")
    print("Колличество файлов для обработки: " + str(count))
    print()
    return count, listpath, dir2


def cols(file):
    wookbook = openpyxl.load_workbook(file)  # открываю исходный эксель файл
    worksheet = wookbook.active
    a = worksheet.max_row
    #b = worksheet.max_column
    i = 0
    collist = []
    for row in range(1, a):
        for col in range(1, 20):
            cell = worksheet.cell(row, col)
            if cell.value == 'Дата' or cell.value == 'Кредит' or cell.value == 'Назначение':
                i += 1
                collist.append(col)
                if i == 3:
                    coldata, colsumma, colost = collist[0], collist[1], collist[2]
                    return coldata, colsumma, colost
    if i < 3:
        print('Искомых наименований столбцов не оказалось. Принтскрин ошибки пришлите на электроннyю почту: ')


def obr():
    tic = time.perf_counter()
    count, listpath, dir2 = pathf()
    for k in range(0, count):
        conch, conm = connnect()
        listch, listm = [], []
        file = listpath[k].split('\\')[-1].strip()  # отсоединяю имя файла для последующего сохранения нового файла
        print(f'Читаю файл № {k+1}: ' + str(file))
        coldata, colsumma, colost = cols(listpath[k])
        wookbook = openpyxl.load_workbook(listpath[k])  # открываю исходный эксель файл
        worksheet = wookbook.active
        a = worksheet.max_row
        # b = worksheet.max_column
        print(f'Колличество строк для обработки в файле № {k+1}: ' + str(a))
        summach, summam = 0, 0  # объявляю переменную для общей суммы
        countm = 0  # счетчик кол-ва лицевых мун.сектора
        countch = 0  # счетчик кол-ва лицевых час.сектора
        for row in range(1, a):
            summa = worksheet.cell(row, colsumma).value  # сумма
            if summa:
                ost = re.findall('(?<![0-9])[0-9]{7,13}(?![0-9])', str((worksheet.cell(row, colost)).value))  # комментарий
                if file.startswith('Statement'):
                    data = (str((worksheet.cell(row, coldata)).value)[:-9]).replace('-', '.')  # дата
                    data = data.split('.')
                    data.reverse()
                    data = '.'.join(data)
                else:
                    data = str((worksheet.cell(row, coldata)).value)  # дата
                chcursor = conch.cursor()  # открываю курсор к частникам
                mcursor = conm.cursor()  # открываю курсор к муниципалке
                for nls in ost:
                    sel = """select nls from _abonent a where a.nls=(?) or a.new_nls=(?);"""  # проверка на наличие лицевых в базах
                    chcursor.execute(sel, nls, nls)  # выполняем запрос с параметрами к частникам
                    mcursor.execute(sel, nls, nls)  # выполняем запрос с параметрами к муниципалке
                    resch = chcursor.fetchone()  # полученный результат проверки лицевого из час.сектора
                    resm = mcursor.fetchone()  # полученный результат проверки лицевого из мун.сектора
                    if resch is None:  # пустоту зануляю для корректного сложения(чисто для новоросса)
                        resch = ['']
                    if resm is None:
                        resm = ['']
                    nls = resch[0].rstrip() + resm[0].rstrip()  # то самое сложение(конкатенация) для финальной зачистки лицевого
                    if nls:  # убираю дубли при записи в фаил
                        summa = round(summa * 100)
                        print('Строка номер:' + str(row) + '. Лицевой: ' + nls + '. Сумма платежа: ' + str(summa/100))
                        if nls.find('2316'):
                            summach += summa
                            listch.append([data, nls, summa])
                            countch += 1
                        elif nls.find('2315'):
                            summam += summa
                            listm.append([data, nls, summa])
                            countm += 1
                        break
                chcursor.close()  # закрываю курсор частников
                mcursor.close()  # закрываю курсор муниципалки
        count -= 1
        conch.commit()  # закрываем подключение
        conm.commit()  # закрываем подключение
        summach = round(summach/100, 2)
        summam = round(summam/100, 2)
        listch.append(summach)
        listch.append(countch)
        listm.append(summam)
        listm.append(countm)
        if len(listm) < 3:
            print('Нечего обрабатывать для муниципалки :(')
        else:
            mlist, mres, mdata = zagm(listm)
            telom(mlist, mres, mdata)
            print(f'Всего обработано лицевых в муниципальном секторе: ' + str(countm) + ' на сумму ' + str(summam))
        if len(listch) < 3:
            print('Нечего обрабатывать для частников :(')
        else:
            chlist, chres, chdata = zagch(listch)
            teloch(chlist, chres, chdata)
            print(f'Всего обработано лицевых в частном секторе: ' + str(countch) + ' на сумму ' + str(summach))
        print(f'Файл № {k + 1} прочитан: ' + str(file))
        os.replace(listpath[k], dir2+file)  # перенос файла
        print(f'Файл № {k + 1} перенесен в {dir2}')
        print()
    print("Осталось считать строк: " + str(count))
    toc = time.perf_counter()
    print(f"Вычисление заняло {toc - tic:0.4f} секунд")
    input('Нажмите Enter для выхода\n')


def reestr():  # функция для полцчения id последнего реестра +1
    conch, conm = connnect()
    chtcur = conch.cursor()
    mtcur = conm.cursor()
    sel = """select max(reestr_id) + 1	from _reestr_zag"""
    chtcur.execute(sel)
    mtcur.execute(sel)
    rescht = chtcur.fetchone()[0]
    resmt = mtcur.fetchone()[0]
    chtcur.close()
    mtcur.close()
    conch.commit()
    conm.commit()
    return rescht, resmt


def period():  # функция получения ОП
    conch, conm = connnect()
    chtcur = conch.cursor()
    mtcur = conm.cursor()
    sel = """select rperiod_beg	from finparams"""
    chtcur.execute(sel)
    mtcur.execute(sel)
    percht = chtcur.fetchone()[0]
    permt = mtcur.fetchone()[0]
    chtcur.close()
    mtcur.close()
    conch.commit()
    conm.commit()
    return percht, permt


def config():  # функция получения конфига
    conch, conm = connnect()
    chtcur = conch.cursor()
    mtcur = conm.cursor()
    sel = """select max(adr) from payerJ_config where pay_id = 100100"""
    chtcur.execute(sel)
    mtcur.execute(sel)
    configcht = chtcur.fetchone()[0]
    configmt = mtcur.fetchone()[0]
    chtcur.close()
    mtcur.close()
    conch.commit()
    conm.commit()
    return configcht, configmt


def zagch(listch):  # функция создания заголовка реестра
    conch, conm = connnect()
    rescht, resmt = reestr()
    percht, permt = period()
    configcht, configmt = config()
    obdatach = listch[0][0]  # дата заголовка частников и 3 строки преобразования
    datech = obdatach.split('.')
    datech[0], datech[2] = datech[2], datech[0]
    obdatach = '-'.join(datech)+' 00:00:00'  # готовая дата заголовка частников
    obsumch = str(listch[-2]).replace('.', ',')
    summaobch = ''
    if obsumch.find(',') == -1:
        summaobch = obsumch + ',00'
    elif obsumch.find(','):
        if obsumch[-2] == ',':
            summaobch = obsumch + '0'
        else:
            summaobch = obsumch
    obsumch = summaobch  # Готовая сумма заголовка частников
    kolvoch = str(listch[-1])  # кол-во платежей частников
    chpaycur = conch.cursor()
    paysel = """SELECT pay_id FROM payerJ WHERE bank=3 AND payment_type = 10"""
    chpaycur.execute(paysel)
    chpayid = chpaycur.fetchone()[0]
    chpaycur.close()
    chusercur = conch.cursor()
    usersel = """SELECT crt_u_id FROM _reestr_zag WHERE payers = (?) AND reestr_id IN (SELECT MAX(reestr_id) FROM _reestr_zag WHERE payers = (?))"""
    chusercur.execute(usersel, chpayid, chpayid)
    chuserid = chusercur.fetchone()[0]
    chusercur.close()
    chtcur = conch.cursor()
    sel = (
        "insert into _reestr_zag(reestr_id, \"type\", \"type_id\", payers, date_bank, kolvo, oplata, comiss, vozmeqenie, peni, \"status\", status_str,\n"
        "	crt_date, crt_u_id, fDel, comment, rperiod_beg, inPaket, one_inp, storno,\n"
        "	adr_create, adr_server,credit_rest_comiss, payerj_config_adr)\n"
        "	select ? reestr_id,\n"
        "	0 \"type\",\n"
        "	5 \"type_id\",\n"
        "	? payers,\n"
        "	? oplata_date, -- дата текущего реестра\n"
        "	? kolvo,\n"
        "	? oplata,	-- сумма платежа в копейках\n"
        "	0 comiss,\n"
        "	0 vozmeqenie, -- та же сумма платежа\n"
        "	0 peni,\n"
        "	2 \"status\",\n"
        "	12 status_str,\n"
        "	? crt_date, -- текущая дата\n"
        "	? crt_u_id,\n"
        "	0 fDel,\n"
        "	'1' comment, -- составляется в соответствии с данными исходного реестра\n"
        "	? rperiod_beg, -- текущий ОП\n"
        "	0 inPaket,\n"
        "	1 one_inp,\n"
        "	0 storno,\n"
        "	? adr_create, \n"
        "	? adr_server,\n"
        "	0 credit_rest_comiss,\n"
        "	? payerj_config_adr")
    chtcur.execute(sel, rescht, chpayid, obdatach, kolvoch, obsumch, obdatach, chuserid, percht, rescht, rescht, configcht)
    chtcur.close()
    conch.commit()
    chlist = listch
    chres = rescht
    chdata = obdatach
    return chlist, chres, chdata


def zagm(listm):  # функция создания заголовка реестра
    conch, conm = connnect()
    rescht, resmt = reestr()
    percht, permt = period()
    configcht, configmt = config()
    obdatam = listm[0][0]  # дата заголовка муниципалки и 3 строки преобразования
    datem = obdatam.split('.')
    datem[0], datem[2] = datem[2], datem[0]
    obdatam = '-'.join(datem) + ' 00:00:00'  # готовая дата заголовка муниципалки
    summaobm = ''
    obsumm = str(listm[-2]).replace('.', ',')
    if obsumm.find(',') == -1:
        summaobm = obsumm + ',00'
    elif obsumm.find(','):
        if obsumm[-2] == ',':
            summaobm = obsumm + '0'
        else:
            summaobm = obsumm
    obsumm = summaobm  # Готовая сумма заголовка муниципалки
    kolvom = str(listm[-1])  # кол-во платежей муниципалки
    mpaycur = conm.cursor()
    paysel = """SELECT pay_id FROM payerJ WHERE bank=3 AND payment_type = 10"""
    mpaycur.execute(paysel)
    mpayid = mpaycur.fetchone()[0]
    mpaycur.close()
    musercur = conm.cursor()
    usersel = """SELECT crt_u_id FROM _reestr_zag WHERE payers = (?) AND reestr_id IN (SELECT MAX(reestr_id) FROM _reestr_zag WHERE payers = (?))"""
    musercur.execute(usersel, mpayid, mpayid)
    muserid = musercur.fetchone()[0]
    musercur.close()
    mtcur = conm.cursor()
    sel = (
        "insert into _reestr_zag(reestr_id, \"type\", \"type_id\", payers, date_bank, kolvo, oplata, comiss, vozmeqenie, peni, \"status\", status_str,\n"
        "	crt_date, crt_u_id, fDel, comment, rperiod_beg, inPaket, one_inp, storno,\n"
        "	adr_create, adr_server,credit_rest_comiss, payerj_config_adr)\n"
        "	select ? reestr_id,\n"
        "	0 \"type\",\n"
        "	5 \"type_id\",\n"
        "	? payers,\n"
        "	? oplata_date, -- дата текущего реестра\n"
        "	? kolvo,\n"
        "	? oplata,	-- сумма платежа в копейках\n"
        "	0 comiss,\n"
        "	0 vozmeqenie, -- та же сумма платежа\n"
        "	0 peni,\n"
        "	2 \"status\",\n"
        "	12 status_str,\n"
        "	? crt_date, -- текущая дата\n"
        "	? crt_u_id,\n"
        "	0 fDel,\n"
        "	'1' comment, -- составляется в соответствии с данными исходного реестра\n"
        "	? rperiod_beg, -- текущий ОП\n"
        "	0 inPaket,\n"
        "	1 one_inp,\n"
        "	0 storno,\n"
        "	? adr_create, \n"
        "	? adr_server,\n"
        "	0 credit_rest_comiss,\n"
        "	? payerj_config_adr")
    mtcur.execute(sel, resmt, mpayid, obdatam, kolvom, obsumm, obdatam, muserid, permt, resmt, resmt, configmt)
    mtcur.close()
    conm.commit()
    mlist = listm
    mres = resmt
    mdata = obdatam
    return mlist, mres, mdata


def telom(mlist, mres, mdata):  # функция внутренности реестра
    conch, conm = connnect()
    mnlscur = conm.cursor()
    mtcur = conm.cursor()
    for i in range(len(mlist)-2):
        nlsvremm = mlist[i][1]
        sumvremm = mlist[i][2]
        selnlsm = """select nls_id from _abonent a where a.nls=(?);"""
        mnlscur.execute(selnlsm, nlsvremm)
        nlsidm = mnlscur.fetchone()[0]
        selm = (
            "insert into _reestr_str(reestr, nls_id, date_beg, date_end, summa, fDistrpayment, fInpErr, fDoubleInp, fDelete, reference_type, oplata_group_id, date_plat, date_ko, nls)\n"
            "        select ? reestr,-- номер реестра\n"
            "        ? nls_id, -- Код лс из выборки\n"
            "        ? date_beg, -- дата исходной оплаты\n"
            "        ? date_end,\n"
            "        ? summa, -- размер оплаты в копейках\n"
            "        0 fDistrpayment,\n"
            "        0 fInpErr,\n"
            "        0 fDoubleInp,\n"
            "        0 fDelete,\n"
            "        0 reference_type,\n"
            "        1 oplata_group_id,\n"
            "        ? date_plat, -- дата исходной оплаты\n"
            "        ? date_ko,\n"
            "        ? nls -- №лс из выборки")
        mtcur.execute(selm, mres, nlsidm, mdata, mdata, sumvremm, mdata, mdata, nlsvremm)
    mnlscur.close()
    mtcur.close()
    conm.commit()


def teloch(chlist, chres, chdata):  # функция внутренности реестра
    conch, conm = connnect()
    chnlscur = conch.cursor()
    chtcur = conch.cursor()
    for i in range(len(chlist)-2):
        nlsvrem = chlist[i][1]
        sumvrem = chlist[i][2]
        selnls = """select nls_id from _abonent a where a.nls=(?);"""
        chnlscur.execute(selnls, nlsvrem)
        nlsid = chnlscur.fetchone()[0]
        sel = (
            "insert into _reestr_str(reestr, nls_id, date_beg, date_end, summa, fDistrpayment, fInpErr, fDoubleInp, fDelete, reference_type, oplata_group_id, date_plat, date_ko, nls)\n"
            "        select ? reestr,-- номер реестра\n"
            "        ? nls_id, -- Код лс из выборки\n"
            "        ? date_beg, -- дата исходной оплаты\n"
            "        ? date_end,\n"
            "        ? summa, -- размер оплаты в копейках\n"
            "        0 fDistrpayment,\n"
            "        0 fInpErr,\n"
            "        0 fDoubleInp,\n"
            "        0 fDelete,\n"
            "        0 reference_type,\n"
            "        1 oplata_group_id,\n"
            "        ? date_plat, -- дата исходной оплаты\n"
            "        ? date_ko,\n"
            "        ? nls -- №лс из выборки")
        chtcur.execute(sel, chres, nlsid, chdata, chdata, sumvrem, chdata, chdata, nlsvrem)
    chnlscur.close()
    chtcur.close()
    conch.commit()


obr()
